import json
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import config

import logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


def load_analyzed_leads(path: str) -> list[dict]:
    """Load analyzed leads from JSON."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def format_social_links(profiles: dict) -> str:
    """Format social profile URLs into a readable string."""
    links = []
    for platform, url in profiles.items():
        if url:
            links.append(f"{platform}: {url}")
    return "\n".join(links) if links else "None found"


def leads_to_dataframe(leads: list[dict]) -> pd.DataFrame:
    """Convert analyzed leads to a DataFrame matching the spec columns."""
    rows = []
    for lead in leads:
        row = {
            "Name": lead.get("name", ""),
            "Title": lead.get("title", ""),
            "Company": lead.get("company", ""),
            "Industry": lead.get("industry", ""),
            "LinkedIn URL": lead.get("linkedin_url", ""),
            "Social Media Links": format_social_links(lead.get("social_profiles", {})),
            "Tone Profile Summary": lead.get("tone_profile", "Not analyzed"),
            "Suggested Connection Message": lead.get("connection_message", ""),
            "Suggested Follow-up Message": lead.get("followup_message", ""),
            "Key Interests": ", ".join(lead.get("key_interests", [])),
            "Talking Points": "\n".join(f"- {tp}" for tp in lead.get("talking_points", [])),
            "Data Quality": lead.get("data_quality", "unknown"),
            "Notes": lead.get("notes", ""),
        }
        rows.append(row)
    return pd.DataFrame(rows)


def style_workbook(filepath: str):
    """Apply formatting to the Excel workbook."""
    wb = load_workbook(filepath)
    ws = wb.active

    # Header style - green per user preference
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D8B2D", end_color="2D8B2D", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    column_widths = {
        "A": 20, "B": 18, "C": 25, "D": 15, "E": 35,
        "F": 30, "G": 40, "H": 45, "I": 45,
        "J": 30, "K": 40, "L": 15, "M": 30,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(filepath)


def write_excel():
    """Main function: load analyzed leads, write formatted Excel."""
    analyzed_path = config.ANALYZED_LEADS_PATH
    if not os.path.exists(analyzed_path):
        logger.error(f"Analyzed leads file not found: {analyzed_path}")
        logger.info("Run Phase 2 (Claude Code analysis) first.")
        return

    leads = load_analyzed_leads(analyzed_path)
    logger.info(f"Loaded {len(leads)} analyzed leads")

    df = leads_to_dataframe(leads)

    # Output file
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(config.OUTPUT_DIR, f"heyva_leads_{timestamp}.xlsx")

    df.to_excel(output_path, index=False, engine="openpyxl")
    style_workbook(output_path)

    logger.info(f"Excel file saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    write_excel()
